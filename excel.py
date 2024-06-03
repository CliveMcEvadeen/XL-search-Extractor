
from openpyxl import Workbook
from openpyxl import worksheet
from openpyxl.styles import Font,Alignment,Color,PatternFill
import os

desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

# Create a new workbook
wb = Workbook()
ws = wb.active  # Get the active worksheet

# Define headers for student information
ws.cell(row=2, column=2).value = "NAME"
ws.cell(row=2, column=3).value = "LIN"

# Define example student data
student_data = [
    ("John Doe", "12345"),
    ("Jane Smith", "54321"),
]

# Write student data
row_num = 3
for name, student_id in student_data:
    ws.cell(row=row_num, column=2).value = name
    ws.cell(row=row_num, column=3).value = student_id
    row_num += 1  # Move to next row for next student

# Define headers for the first subject (replace with actual subject names)
subject_name1 = "Subject 1"
ws.cell(row=1, column=4).value = subject_name1
ws.cell(row=2, column=4).value = "C1"
ws.cell(row=2, column=5).value = "C2"
ws.cell(row=2, column=6).value = "C3"
ws.cell(row=2, column=7).value = "Final Exam"

header_font = Font(bold=True)
header_align = Alignment(horizontal='center', vertical='center')
data_align = Alignment(horizontal='center', vertical='center')
header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')



# Define headers for the second subject (replace with actual subject names)
subject_name2 = "Subject 2"
ws.cell(row=1, column=8).value = subject_name2
ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=11)
ws.cell(row=1, column=8).alignment=data_align
ws.cell(row=1, column=8).fill=header_fill
ws.cell(row=1, column=8).font=header_font
ws.cell(row=1, column=8).fill=header_fill

# setting column width to auto


ws.cell(row=1, column=8).value = subject_name2
ws.cell(row=2, column=8).value = "C1"
ws.cell(row=2, column=9).value = "C2"
ws.cell(row=2, column=10).value = "C3"
ws.cell(row=2, column=11).value = "Final Exam"

# Add some example marks (replace with actual data import logic)
example_marks = [
    [50, 75, 80, 90],  # John Doe's marks
    [85, 60, 95, 78],  # Jane Smith's marks
]

row_num = 3
for marks in example_marks:
    for col_num, mark in enumerate(marks, start=4):  # Start from column D
        ws.cell(row=row_num, column=col_num).value = mark
    row_num += 1

# Save the workbook as a spreadsheet file
wb.save(os.path.join(desktop_path, "student_marks2.xlsx"))

print("Spreadsheet created successfully and saved to your Desktop!")