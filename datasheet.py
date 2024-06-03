import os
import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def generate_data_sheet():
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
    folder_name = "datasheets"
    folder_path = os.path.join(desktop_path, folder_name)
    
    # Create the folder if it doesn't exist
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    # Generate a random number for the filename
    random_number = random.randint(1000, 9999)
    file_name = f"datasheet_{random_number}.xlsx"
    file_path = os.path.join(folder_path, file_name)

    # Create a new workbook
    wb = Workbook()
    ws = wb.active  # Get the active worksheet

    # Define title for the sheet
    title = "Student Marks Report"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=42)  # Merge cells from A1 to AN1
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Define headers for student information
    ws.cell(row=3, column=1).value = "NAME"
    ws.cell(row=3, column=2).value = "LIN"

    header_font = Font(bold=True)
    header_align = Alignment(horizontal='center', vertical='center')
    data_align = Alignment(horizontal='center', vertical='center')

    # Define colors for subjects
    colors = ['FFC000', '00B0F0']

    for i in range(1, 11):
        subject_name = f"Subject {i}"
        start_col = 3 + (i - 1) * 4
        end_col = start_col + 3
        
        # Set alternating colors for subjects
        header_fill = PatternFill(start_color=colors[i % 2], end_color=colors[i % 2], fill_type='solid')
        
        # Merge cells for subject name and set styles
        ws.cell(row=2, column=start_col).value = subject_name
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        ws.cell(row=2, column=start_col).alignment = header_align
        ws.cell(row=2, column=start_col).fill = header_fill
        ws.cell(row=2, column=start_col).font = header_font

        # Set headers for each component of the subject
        ws.cell(row=3, column=start_col).value = "C1"
        ws.cell(row=3, column=start_col + 1).value = "C2"
        ws.cell(row=3, column=start_col + 2).value = "C3"
        ws.cell(row=3, column=start_col + 3).value = "FS"
        for col in range(start_col, end_col + 1):
            ws.cell(row=3, column=col).alignment = header_align
            ws.cell(row=3, column=col).fill = header_fill
            ws.cell(row=3, column=col).font = header_font

    # Save the workbook as a spreadsheet file
    wb.save(file_path)

    print(f"Spreadsheet created successfully and saved to {file_path}")

generate_data_sheet()




